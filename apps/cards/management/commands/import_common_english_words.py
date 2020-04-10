import openpyxl
from django.core.files import File
from django.core.management import BaseCommand

from apps.cards.models import Card

ROOT_DICTIONARY_FOLDER = './oxford_dictionary_scraper/'


class Command(BaseCommand):
    def handle(self, *args, **options):
        word_dictionary = openpyxl.load_workbook(
            filename=f'{ROOT_DICTIONARY_FOLDER}oxford_dictionary_3000_words.xlsx'
        )
        sheet = word_dictionary.worksheets[0]
        for i in range(1, sheet.max_row):
            text = sheet.cell(i, 1).value
            sound_path = sheet.cell(i, 3).value

            if sound_path is None:
                continue

            f = open(f'{ROOT_DICTIONARY_FOLDER}{sound_path}', 'rb')

            sound = File(f)
            filename = sound_path.split('/')
            card = Card(text=text)
            card.sound.save(filename[1], sound)
            card.save()
